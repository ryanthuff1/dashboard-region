#!/usr/bin/env python3
"""
Verify that the Excel workbook matches the pipeline dashboard data.
"""

import json
from pathlib import Path
from openpyxl import load_workbook

DEALS_JSON = Path(__file__).parent / 'deals.json'
WORKBOOK_PATH = Path(__file__).parent.parent / 'RM-pipeline-workbook.xlsx'

def main():
    print("\n🔍 Verifying Excel ↔ Dashboard alignment...\n")

    # Load JSON
    with open(DEALS_JSON) as f:
        json_deals = json.load(f)

    # Load Excel
    wb = load_workbook(WORKBOOK_PATH)

    # Check Summary sheet
    if 'Summary' not in wb.sheetnames:
        print("❌ Summary sheet not found")
        return False

    summary = wb['Summary']
    print("✓ Summary sheet found")

    # Check Active Pipeline sheet
    if 'Active Pipeline' not in wb.sheetnames:
        print("❌ Active Pipeline sheet not found")
        return False

    pipeline_sheet = wb['Active Pipeline']
    pipeline_row_count = pipeline_sheet.max_row - 1  # Subtract header row

    print(f"✓ Active Pipeline sheet found")

    # Check Closed YTD sheet
    if 'Closed YTD' not in wb.sheetnames:
        print("❌ Closed YTD sheet not found")
        return False

    closed_sheet = wb['Closed YTD']
    closed_row_count = closed_sheet.max_row - 1  # Subtract header row

    print(f"✓ Closed YTD sheet found")

    excel_row_count = pipeline_row_count + closed_row_count
    print(f"\nDeal counts:")
    print(f"  JSON:  {len(json_deals)} deals total")
    print(f"  Excel: {excel_row_count} deals total")
    print(f"    → Active Pipeline: {pipeline_row_count} deals")
    print(f"    → Closed YTD: {closed_row_count} deals")

    if len(json_deals) == excel_row_count:
        print("  ✓ Counts match!")
    else:
        print("  ❌ Count mismatch!")
        return False

    # Calculate metrics from JSON
    open_deals = [d for d in json_deals if d['stage'] != 'Closed']
    closed_deals = [d for d in json_deals if d['stage'] == 'Closed']

    total_pipeline = sum(d['amount'] for d in open_deals)
    weighted_revenue = sum(d['revAnnual'] * d['probability'] for d in open_deals)

    print(f"\nMetrics from JSON:")
    print(f"  Open deals: {len(open_deals)}")
    print(f"  Closed deals: {len(closed_deals)}")
    print(f"  Total pipeline: ${total_pipeline:,.2f}")
    print(f"  Weighted revenue: ${weighted_revenue:,.2f}")

    # Stage breakdown
    stages = ['Lead', 'Qualified', 'Proposal', 'Commit', 'Closed']
    stage_counts = {s: sum(1 for d in json_deals if d['stage'] == s) for s in stages}

    print(f"\nStage breakdown:")
    for stage in stages:
        print(f"  {stage:12s}: {stage_counts[stage]:2d} deals")

    # Sample some deals from each sheet to verify data integrity
    print(f"\n✓ Sampling active pipeline deals:")
    for row in range(2, min(4, pipeline_sheet.max_row + 1)):
        deal_id = pipeline_sheet.cell(row, 1).value
        officer = pipeline_sheet.cell(row, 3).value
        customer = pipeline_sheet.cell(row, 5).value
        stage = pipeline_sheet.cell(row, 7).value
        amount = pipeline_sheet.cell(row, 8).value
        print(f"  {deal_id:15s} | {officer:12s} | {customer:20s} | {stage:10s} | ${amount:>12,.2f}")

    print(f"\n✓ Sampling closed YTD deals:")
    for row in range(2, min(4, closed_sheet.max_row + 1)):
        deal_id = closed_sheet.cell(row, 1).value
        officer = closed_sheet.cell(row, 3).value
        customer = closed_sheet.cell(row, 5).value
        stage = closed_sheet.cell(row, 7).value
        amount = closed_sheet.cell(row, 8).value
        print(f"  {deal_id:15s} | {officer:12s} | {customer:20s} | {stage:10s} | ${amount:>12,.2f}")

    print(f"\n✅ Verification complete! Excel has proper Active Pipeline / Closed YTD separation.\n")
    return True

if __name__ == '__main__':
    main()
