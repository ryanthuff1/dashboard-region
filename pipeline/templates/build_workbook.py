"""
Build the RM pipeline workbook.

Run from the repo root:
    python3 pipeline/templates/build_workbook.py

Produces: pipeline/templates/RM-pipeline-template.xlsx

Design (per agreed spec):
  - One tab per RM, named with their initial+last name
  - Probability auto-fills from Stage but is overridable (RM types over the formula)
  - Days in Stage auto-calculates from "Date Entered Stage" via TODAY()
  - Hidden Lookup tab drives dropdowns and lookup formulas
  - Summary tab rolls up each RM's totals
  - README tab explains how to fill it out
  - 30 pre-formatted empty rows on each RM tab with validation + formulas baked in
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT_PATH = Path(__file__).parent / "RM-pipeline-template.xlsx"

# Roster (matches pipeline/js/data.js)
OFFICERS = [
    {"name": "M. Reyes",    "market": "Houston"},
    {"name": "D. Patel",    "market": "Houston"},
    {"name": "A. Tran",     "market": "Houston"},
    {"name": "K. Williams", "market": "Houston"},
    {"name": "R. Chen",     "market": "Dallas"},
    {"name": "S. Garcia",   "market": "Dallas"},
    {"name": "B. Foster",   "market": "Dallas"},
]

PRODUCTS = ["Comm Loan", "Deposits", "Treasury Mgmt", "Other Fees"]
STAGES   = ["Lead", "Qualified", "Proposal", "Commit", "Closed"]
STAGE_PROB    = {"Lead": 0.10, "Qualified": 0.25, "Proposal": 0.50, "Commit": 0.80, "Closed": 1.00}
PRODUCT_RATIO = {"Comm Loan": 0.034, "Deposits": 0.005, "Treasury Mgmt": 1.0, "Other Fees": 1.0}

DEAL_ROWS = 30  # blank rows pre-formatted per RM tab

# ── Styles ─────────────────────────────────────────────────────────
HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="1864AB")
SUBTLE_FILL  = PatternFill("solid", fgColor="F1F3F5")
CALC_FILL    = PatternFill("solid", fgColor="E7F5FF")  # auto-calculated cells
TITLE_FONT   = Font(name="Calibri", size=16, bold=True, color="1864AB")
LABEL_FONT   = Font(name="Calibri", size=11, bold=True)
THIN         = Side(border_style="thin", color="DEE2E6")
BOX_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RED_FILL     = PatternFill("solid", fgColor="FFE3E3")
AMBER_FILL   = PatternFill("solid", fgColor="FFF3BF")

wb = Workbook()
# Remove default sheet — we'll add ours in order
wb.remove(wb.active)


# ── README ─────────────────────────────────────────────────────────
def build_readme():
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    ws["A1"] = "RM Pipeline Template"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    rows = [
        ("", ""),
        ("How to use",   "Each RM has their own tab (named with your initial + last name). Fill in only your tab. Submit the workbook when ready and the dashboard will refresh."),
        ("Tabs",         "README · Summary (auto-calculated rollup) · one tab per RM. Lookup is hidden — leave it alone."),
        ("", ""),
        ("Required",     "Customer · Product · Stage · Amount · Date Entered Stage"),
        ("Auto-filled",  "Probability (from Stage), Days in Stage (from date), Annual Rev (from Product+Amount), Weighted Rev (Annual × Probability). Cells with these formulas are light blue."),
        ("Overriding",   "Probability auto-fills from Stage, but you can type a number over the formula if your judgment differs."),
        ("", ""),
        ("Products",     "Comm Loan · Deposits · Treasury Mgmt · Other Fees"),
        ("Stages (prob)", "Lead (10%) · Qualified (25%) · Proposal (50%) · Commit (80%) · Closed (100%)"),
        ("", ""),
        ("Amount units", "Comm Loan / Deposits = loan or deposit balance in dollars. Treasury Mgmt / Other Fees = annual fee revenue in dollars."),
        ("Annual rev",   "For loans/deposits, Annual Rev = Amount × spread (3.4% commercial loan, 0.5% deposit). For TM/Other, Annual Rev = Amount as entered."),
        ("Closed deals", "Mark Stage = Closed when funded/booked. Closed deals count toward Closed YTD on the dashboard but are excluded from open pipeline."),
        ("", ""),
        ("Aging colors", "Days in Stage > 90 highlight red; 61-90 amber. Use these to triage stuck deals."),
        ("Notes column", "Free text. Use it for things the columns don't capture (credit conditions, RM commentary, blockers)."),
    ]
    for i, (label, body) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        ws.cell(row=i, column=1).alignment = Alignment(vertical="top")
        ws.cell(row=i, column=2, value=body).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30 if body else 8


# ── Lookup (hidden) ────────────────────────────────────────────────
def build_lookup():
    ws = wb.create_sheet("Lookup")
    ws["A1"] = "Product"
    ws["B1"] = "Spread"
    for i, p in enumerate(PRODUCTS, start=2):
        ws.cell(row=i, column=1, value=p)
        ws.cell(row=i, column=2, value=PRODUCT_RATIO[p])

    ws["D1"] = "Stage"
    ws["E1"] = "Probability"
    for i, s in enumerate(STAGES, start=2):
        ws.cell(row=i, column=1 + 3, value=s)
        ws.cell(row=i, column=2 + 3, value=STAGE_PROB[s])

    ws["G1"] = "RMs"
    ws["H1"] = "Market"
    for i, o in enumerate(OFFICERS, start=2):
        ws.cell(row=i, column=7, value=o["name"])
        ws.cell(row=i, column=8, value=o["market"])

    # Header styling
    for col_letter in ["A", "B", "D", "E", "G", "H"]:
        c = ws[f"{col_letter}1"]
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for col in ["A", "B", "D", "E", "G", "H"]:
        ws.column_dimensions[col].width = 14

    ws.sheet_state = "hidden"


# ── Per-RM tabs ────────────────────────────────────────────────────
RM_HEADERS = [
    "Customer", "Product", "Stage", "Amount ($)",
    "Probability", "Date Entered Stage", "Days in Stage",
    "Expected Close", "Annual Rev ($)", "Weighted Rev ($)", "Notes",
]
RM_COL_WIDTHS = [26, 16, 14, 14, 12, 18, 12, 14, 16, 16, 40]

# Column letters
COL = {h: get_column_letter(i+1) for i, h in enumerate(RM_HEADERS)}
LAST_DATA_ROW = 1 + DEAL_ROWS  # row 31 if DEAL_ROWS=30

def build_rm_tab(officer):
    name = officer["name"]
    ws = wb.create_sheet(name)

    # Header banner
    ws["A1"] = f"{name} · {officer['market']} · Pipeline"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:K1")
    ws.row_dimensions[1].height = 22

    # Column headers on row 2
    for i, h in enumerate(RM_HEADERS, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX_BORDER
    for i, w in enumerate(RM_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 22

    # Fill formulas + formats for rows 3 .. LAST_DATA_ROW+1
    for r in range(3, LAST_DATA_ROW + 2):
        # Probability (E): VLOOKUP stage on Lookup, but blank if stage blank
        ws.cell(row=r, column=5,
                value=f'=IF(C{r}="","",VLOOKUP(C{r},Lookup!$D$2:$E$6,2,FALSE))')
        ws.cell(row=r, column=5).fill = CALC_FILL
        ws.cell(row=r, column=5).number_format = "0%"

        # Days in Stage (G): TODAY() - Date Entered Stage
        ws.cell(row=r, column=7,
                value=f'=IF(F{r}="","",TODAY()-F{r})')
        ws.cell(row=r, column=7).fill = CALC_FILL
        ws.cell(row=r, column=7).number_format = "0"

        # Annual Rev (I): amount × product spread
        ws.cell(row=r, column=9,
                value=f'=IF(OR(B{r}="",D{r}=""),"",D{r}*VLOOKUP(B{r},Lookup!$A$2:$B$5,2,FALSE))')
        ws.cell(row=r, column=9).fill = CALC_FILL
        ws.cell(row=r, column=9).number_format = '"$"#,##0'

        # Weighted Rev (J): annual rev × probability
        ws.cell(row=r, column=10,
                value=f'=IF(OR(I{r}="",E{r}=""),"",I{r}*E{r})')
        ws.cell(row=r, column=10).fill = CALC_FILL
        ws.cell(row=r, column=10).number_format = '"$"#,##0'

        # Number/date formats for inputs
        ws.cell(row=r, column=4).number_format = '"$"#,##0'         # Amount
        ws.cell(row=r, column=6).number_format = "m/d/yyyy"         # Date Entered Stage
        ws.cell(row=r, column=8).number_format = "m/d/yyyy"         # Expected Close

        # Light row banding for readability
        if r % 2 == 0:
            for col_idx in range(1, 12):
                cell = ws.cell(row=r, column=col_idx)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = SUBTLE_FILL

    # Data validation — Product dropdown
    dv_product = DataValidation(type="list", formula1="=Lookup!$A$2:$A$5", allow_blank=True,
                                showErrorMessage=True, errorTitle="Invalid product",
                                error="Pick one of: Comm Loan, Deposits, Treasury Mgmt, Other Fees")
    dv_product.add(f"B3:B{LAST_DATA_ROW+1}")
    ws.add_data_validation(dv_product)

    # Stage dropdown
    dv_stage = DataValidation(type="list", formula1="=Lookup!$D$2:$D$6", allow_blank=True,
                              showErrorMessage=True, errorTitle="Invalid stage",
                              error="Pick one of: Lead, Qualified, Proposal, Commit, Closed")
    dv_stage.add(f"C3:C{LAST_DATA_ROW+1}")
    ws.add_data_validation(dv_stage)

    # Amount must be >= 0
    dv_amount = DataValidation(type="decimal", operator="greaterThanOrEqual",
                               formula1=0, allow_blank=True)
    dv_amount.add(f"D3:D{LAST_DATA_ROW+1}")
    ws.add_data_validation(dv_amount)

    # Probability 0-1
    dv_prob = DataValidation(type="decimal", operator="between",
                             formula1=0, formula2=1, allow_blank=True,
                             errorTitle="Invalid probability",
                             error="Enter a decimal between 0 and 1 (e.g. 0.25 for 25%)")
    dv_prob.add(f"E3:E{LAST_DATA_ROW+1}")
    ws.add_data_validation(dv_prob)

    # Date validators
    for col_letter, name_ in [("F", "Date Entered Stage"), ("H", "Expected Close")]:
        dv_date = DataValidation(type="date", allow_blank=True,
                                 errorTitle=f"Invalid date — {name_}",
                                 error="Enter a date (mm/dd/yyyy)")
        dv_date.add(f"{col_letter}3:{col_letter}{LAST_DATA_ROW+1}")
        ws.add_data_validation(dv_date)

    # Conditional formatting: Days in Stage
    days_range = f"G3:G{LAST_DATA_ROW+1}"
    ws.conditional_formatting.add(
        days_range,
        CellIsRule(operator="greaterThan", formula=["90"], fill=RED_FILL,
                   font=Font(color="9F1239", bold=True)),
    )
    ws.conditional_formatting.add(
        days_range,
        CellIsRule(operator="between", formula=["61", "90"], fill=AMBER_FILL,
                   font=Font(color="92400E", bold=True)),
    )

    # Freeze header
    ws.freeze_panes = "A3"


# ── Summary ────────────────────────────────────────────────────────
def build_summary():
    ws = wb.create_sheet("Summary", 1)  # second tab, right after README

    ws["A1"] = "Pipeline Summary (auto-calculated)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 22

    headers = ["RM", "Market", "# Open Deals", "Total Balance ($)",
               "Annual Rev ($)", "Weighted Rev ($)", "Closed YTD ($)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

    widths = [22, 14, 16, 18, 16, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, o in enumerate(OFFICERS, start=3):
        sheet = f"'{o['name']}'"
        rng_stage  = f"{sheet}!C3:C{LAST_DATA_ROW+1}"
        rng_amount = f"{sheet}!D3:D{LAST_DATA_ROW+1}"
        rng_annual = f"{sheet}!I3:I{LAST_DATA_ROW+1}"
        rng_weight = f"{sheet}!J3:J{LAST_DATA_ROW+1}"

        ws.cell(row=i, column=1, value=o["name"]).font = LABEL_FONT
        ws.cell(row=i, column=2, value=o["market"])
        # Count open (stage <> "" and <> "Closed")
        ws.cell(row=i, column=3,
                value=f'=COUNTIFS({rng_stage},"<>",{rng_stage},"<>Closed")')
        ws.cell(row=i, column=4,
                value=f'=SUMIFS({rng_amount},{rng_stage},"<>Closed",{rng_stage},"<>")')
        ws.cell(row=i, column=5,
                value=f'=SUMIFS({rng_annual},{rng_stage},"<>Closed",{rng_stage},"<>")')
        ws.cell(row=i, column=6,
                value=f'=SUMIFS({rng_weight},{rng_stage},"<>Closed",{rng_stage},"<>")')
        ws.cell(row=i, column=7,
                value=f'=SUMIFS({rng_annual},{rng_stage},"Closed")')

        ws.cell(row=i, column=4).number_format = '"$"#,##0'
        ws.cell(row=i, column=5).number_format = '"$"#,##0'
        ws.cell(row=i, column=6).number_format = '"$"#,##0'
        ws.cell(row=i, column=7).number_format = '"$"#,##0'

    # Totals row
    total_row = 3 + len(OFFICERS)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="212529")
    ws.cell(row=total_row, column=2).fill = PatternFill("solid", fgColor="212529")
    for col in range(3, 8):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}3:{col_letter}{total_row - 1})"
        c = ws.cell(row=total_row, column=col, value=formula)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E9ECEF")
        if col >= 4:
            c.number_format = '"$"#,##0'

    ws.freeze_panes = "A3"


# ── Build in tab order: README, Summary, RM tabs, (hidden Lookup) ──
build_readme()
build_summary()
build_lookup()
for o in OFFICERS:
    build_rm_tab(o)

# Move sheets into desired order: README, Summary, RM tabs..., Lookup (hidden, last)
ordered = ["README", "Summary"] + [o["name"] for o in OFFICERS] + ["Lookup"]
wb._sheets = [wb[name] for name in ordered]

# Set the active tab to README on open
wb.active = 0

OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}  ({OUT_PATH.stat().st_size:,} bytes)")
