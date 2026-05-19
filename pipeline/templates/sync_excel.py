#!/usr/bin/env python3
"""
Sync the Excel workbook with the pipeline dashboard's synthetic deals.
This ensures the Excel template matches exactly what the web dashboard displays.

Usage:
    python3 sync_excel.py
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load the deals JSON (generated from dump_deals.js)
DEALS_JSON = Path(__file__).parent / 'deals.json'
WORKBOOK_PATH = Path(__file__).parent.parent / 'RM-pipeline-workbook.xlsx'

def load_deals():
    """Load deals from JSON file."""
    if not DEALS_JSON.exists():
        print(f"❌ Error: {DEALS_JSON} not found")
        print("   Run: node pipeline/templates/dump_deals.js > pipeline/templates/deals.json")
        sys.exit(1)

    with open(DEALS_JSON) as f:
        deals = json.load(f)

    print(f"✓ Loaded {len(deals)} deals from {DEALS_JSON.name}")
    return deals

def format_currency(amount):
    """Format amount as currency."""
    return f"${amount:,.2f}"

def format_percent(prob):
    """Format probability as percentage."""
    return f"{prob:.0%}"

def format_date(date_str):
    """Convert ISO date string to datetime object."""
    return datetime.strptime(date_str, '%Y-%m-%d')

def populate_pipeline_sheet(wb, deals):
    """Populate Active Pipeline sheet with open deals only."""

    # Filter to open deals only
    open_deals = [d for d in deals if d['stage'] != 'Closed']

    # Remove old sheet if exists
    if 'Active Pipeline' in wb.sheetnames:
        del wb['Active Pipeline']

    ws = wb.create_sheet('Active Pipeline', 1)

    # Define headers
    headers = [
        'Deal ID', 'Officer ID', 'Officer Name', 'Market', 'Customer',
        'Product', 'Stage', 'Amount', 'Probability', 'Days in Stage',
        'Age (Days)', 'Expected Close', 'Annual Revenue'
    ]

    # Style for headers
    header_fill = PatternFill(start_color='1864ab', end_color='1864ab', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        bottom=Side(style='thin', color='dee2e6')
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

    # Officer name mapping
    officer_names = {
        'h-reyes': 'M. Reyes',
        'h-patel': 'D. Patel',
        'h-tran': 'A. Tran',
        'h-williams': 'K. Williams',
        'd-chen': 'R. Chen',
        'd-garcia': 'S. Garcia',
        'd-foster': 'B. Foster',
    }

    # Product label mapping
    product_labels = {
        'CommLoan': 'Commercial Loan',
        'Deposits': 'Deposits',
        'TM': 'Treasury Management',
        'Other': 'Other Fees',
    }

    # Write deal data (open deals only)
    row = 2
    for deal in open_deals:
        ws.cell(row, 1, deal['id'])
        ws.cell(row, 2, deal['officerId'])
        ws.cell(row, 3, officer_names.get(deal['officerId'], deal['officerId']))
        ws.cell(row, 4, deal['market'])
        ws.cell(row, 5, deal['customer'])
        ws.cell(row, 6, product_labels.get(deal['product'], deal['product']))
        ws.cell(row, 7, deal['stage'])

        # Amount as number (not string)
        amount_cell = ws.cell(row, 8, deal['amount'])
        amount_cell.number_format = '$#,##0.00'

        # Probability as percentage
        prob_cell = ws.cell(row, 9, deal['probability'])
        prob_cell.number_format = '0%'

        ws.cell(row, 10, deal['daysInStage'])
        ws.cell(row, 11, deal['ageDays'])

        # Expected close as date
        close_cell = ws.cell(row, 12, format_date(deal['expectedClose']))
        close_cell.number_format = 'mm/dd/yyyy'

        # Annual revenue as number
        rev_cell = ws.cell(row, 13, deal['revAnnual'])
        rev_cell.number_format = '$#,##0.00'

        row += 1

    # Adjust column widths
    column_widths = {
        'A': 14,  # Deal ID
        'B': 12,  # Officer ID
        'C': 14,  # Officer Name
        'D': 10,  # Market
        'E': 20,  # Customer
        'F': 18,  # Product
        'G': 12,  # Stage
        'H': 14,  # Amount
        'I': 12,  # Probability
        'J': 14,  # Days in Stage
        'K': 12,  # Age
        'L': 14,  # Expected Close
        'M': 16,  # Annual Revenue
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    print(f"✓ Populated {len(open_deals)} open deals in 'Active Pipeline' sheet")
    return ws

def populate_closed_sheet(wb, deals):
    """Populate Closed YTD sheet with closed deals."""

    # Filter to closed deals only
    closed_deals = [d for d in deals if d['stage'] == 'Closed']

    # Remove old sheet if exists
    if 'Closed YTD' in wb.sheetnames:
        del wb['Closed YTD']

    ws = wb.create_sheet('Closed YTD', 2)

    # Define headers (same as pipeline sheet)
    headers = [
        'Deal ID', 'Officer ID', 'Officer Name', 'Market', 'Customer',
        'Product', 'Stage', 'Amount', 'Probability', 'Days in Stage',
        'Age (Days)', 'Close Date', 'Annual Revenue'
    ]

    # Style for headers (green theme for closed)
    header_fill = PatternFill(start_color='2b8a3e', end_color='2b8a3e', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(bottom=Side(style='thin', color='dee2e6'))

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

    # Officer name mapping
    officer_names = {
        'h-reyes': 'M. Reyes',
        'h-patel': 'D. Patel',
        'h-tran': 'A. Tran',
        'h-williams': 'K. Williams',
        'd-chen': 'R. Chen',
        'd-garcia': 'S. Garcia',
        'd-foster': 'B. Foster',
    }

    # Product label mapping
    product_labels = {
        'CommLoan': 'Commercial Loan',
        'Deposits': 'Deposits',
        'TM': 'Treasury Management',
        'Other': 'Other Fees',
    }

    # Write closed deal data
    row = 2
    for deal in closed_deals:
        ws.cell(row, 1, deal['id'])
        ws.cell(row, 2, deal['officerId'])
        ws.cell(row, 3, officer_names.get(deal['officerId'], deal['officerId']))
        ws.cell(row, 4, deal['market'])
        ws.cell(row, 5, deal['customer'])
        ws.cell(row, 6, product_labels.get(deal['product'], deal['product']))
        ws.cell(row, 7, deal['stage'])

        # Amount as number
        amount_cell = ws.cell(row, 8, deal['amount'])
        amount_cell.number_format = '$#,##0.00'

        # Probability (always 100% for closed)
        prob_cell = ws.cell(row, 9, 1.0)
        prob_cell.number_format = '0%'

        ws.cell(row, 10, deal['daysInStage'])
        ws.cell(row, 11, deal['ageDays'])

        # Close date
        close_cell = ws.cell(row, 12, format_date(deal['expectedClose']))
        close_cell.number_format = 'mm/dd/yyyy'

        # Annual revenue
        rev_cell = ws.cell(row, 13, deal['revAnnual'])
        rev_cell.number_format = '$#,##0.00'

        row += 1

    # Adjust column widths
    column_widths = {
        'A': 14, 'B': 12, 'C': 14, 'D': 10, 'E': 20,
        'F': 18, 'G': 12, 'H': 14, 'I': 12, 'J': 14,
        'K': 12, 'L': 14, 'M': 16,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    print(f"✓ Populated {len(closed_deals)} closed deals in 'Closed YTD' sheet")
    return ws

def add_summary_sheet(wb, deals):
    """Add a summary sheet with key metrics."""

    # Remove existing summary if it exists
    if 'Summary' in wb.sheetnames:
        del wb['Summary']

    # Create new summary sheet at the beginning
    ws = wb.create_sheet('Summary', 0)

    # Calculate metrics
    open_deals = [d for d in deals if d['stage'] != 'Closed']
    closed_deals = [d for d in deals if d['stage'] == 'Closed']

    total_pipeline = sum(d['amount'] for d in open_deals)
    weighted_pipeline = sum(d['amount'] * d['probability'] for d in open_deals)
    total_revenue = sum(d['revAnnual'] for d in open_deals)
    weighted_revenue = sum(d['revAnnual'] * d['probability'] for d in open_deals)
    closed_revenue = sum(d['revAnnual'] for d in closed_deals)

    # Stage breakdown
    stages = ['Lead', 'Qualified', 'Proposal', 'Commit', 'Closed']
    stage_counts = {s: sum(1 for d in deals if d['stage'] == s) for s in stages}

    # Write summary data
    title_font = Font(bold=True, size=14, color='1864ab')
    header_font = Font(bold=True, size=11)

    ws['A1'] = 'Pipeline Dashboard - Summary'
    ws['A1'].font = title_font

    row = 3
    ws.cell(row, 1, 'Metric').font = header_font
    ws.cell(row, 2, 'Value').font = header_font

    metrics = [
        ('── ACTIVE PIPELINE ──', ''),
        ('Open Deals', len(open_deals)),
        ('Total Pipeline (Balance)', total_pipeline),
        ('Weighted Pipeline', weighted_pipeline),
        ('Total Revenue Potential', total_revenue),
        ('Weighted Revenue', weighted_revenue),
        ('', ''),
        ('── CLOSED YTD ──', ''),
        ('Closed Deals', len(closed_deals)),
        ('Closed Revenue (YTD)', closed_revenue),
        ('', ''),
        ('── STAGE BREAKDOWN ──', ''),
    ]

    row = 4
    for label, value in metrics:
        ws.cell(row, 1, label)
        if isinstance(value, (int, float)) and value > 1000:
            cell = ws.cell(row, 2, value)
            cell.number_format = '$#,##0.00'
        else:
            ws.cell(row, 2, value)
        row += 1

    # Add stage breakdown
    for stage in stages:
        ws.cell(row, 1, f'  {stage}')
        ws.cell(row, 2, stage_counts[stage])
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    print("✓ Added summary sheet")

def main():
    """Main execution."""
    print("\n🔄 Syncing Excel workbook with pipeline dashboard data...\n")

    # Load deals
    deals = load_deals()

    # Load workbook
    if not WORKBOOK_PATH.exists():
        print(f"❌ Error: Workbook not found at {WORKBOOK_PATH}")
        sys.exit(1)

    print(f"✓ Loading workbook: {WORKBOOK_PATH.name}")
    wb = load_workbook(WORKBOOK_PATH)

    # Remove old 'Deals' sheet if it exists (legacy)
    if 'Deals' in wb.sheetnames:
        del wb['Deals']
        print("✓ Removed legacy 'Deals' sheet")

    # Populate sheets
    populate_pipeline_sheet(wb, deals)
    populate_closed_sheet(wb, deals)
    add_summary_sheet(wb, deals)

    # Count splits
    open_count = len([d for d in deals if d['stage'] != 'Closed'])
    closed_count = len([d for d in deals if d['stage'] == 'Closed'])

    # Save workbook
    wb.save(WORKBOOK_PATH)
    print(f"\n✅ Workbook saved: {WORKBOOK_PATH}")
    print(f"   {len(deals)} total deals synced from dashboard")
    print(f"   → {open_count} in 'Active Pipeline'")
    print(f"   → {closed_count} in 'Closed YTD'\n")

if __name__ == '__main__':
    main()
